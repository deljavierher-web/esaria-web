#!/usr/bin/env python3
"""Genera un Excel editable para preparar y registrar llamadas comerciales."""

from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = ROOT / "leads" / "reales" / "leads-reales.json"
DEFAULT_OUTPUT = ROOT / "leads" / "hoja-llamadas-martes.xlsx"
MAIN_SHEET = "Llamadas martes"

CALL_COLUMNS = [
    "Orden",
    "Llamar",
    "Sector",
    "Empresa",
    "Telefono",
    "Contacto",
    "Cargo",
    "Prioridad",
    "Problema probable",
    "Frase de entrada",
    "Estado llamada",
    "Resultado",
    "Fecha cita",
    "Hora cita",
    "Fecha seguimiento",
    "Proximo paso",
    "Notas",
]

EDITABLE_COLUMNS = {
    "Llamar",
    "Estado llamada",
    "Resultado",
    "Fecha cita",
    "Hora cita",
    "Fecha seguimiento",
    "Proximo paso",
    "Notas",
}

STATE_OPTIONS = [
    "Pendiente",
    "Llamado",
    "No responde",
    "Ignorado",
    "No interesa",
    "Interesado",
    "Cita agendada",
    "Seguimiento pendiente",
    "Descartado",
]

RESULT_OPTIONS = [
    "Pendiente",
    "No responde",
    "Contesta recepcion",
    "Hablo con responsable",
    "Pide WhatsApp",
    "Pide email",
    "Agenda cita",
    "Llamar otro dia",
    "No interesa",
    "Numero incorrecto",
]

YES_NO_OPTIONS = ["Si", "No"]

HEADER_FILL = PatternFill("solid", fgColor="0F172A")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BORDER = Border(bottom=Side(style="thin", color="E2E8F0"))


def clean(value: object) -> str:
    return str(value or "").replace("\n", " ").strip()


def read_leads(path: Path, limit: int) -> list[dict[str, object]]:
    if not path.exists():
        raise SystemExit(f"No existe el archivo de leads: {path}")
    leads = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(leads, list):
        raise SystemExit("El JSON de leads debe contener una lista.")

    filtered = [lead for lead in leads if isinstance(lead, dict) and clean(lead.get("telefono"))]
    filtered.sort(key=priority_rank)
    return filtered[:limit]


def priority_rank(lead: dict[str, object]) -> tuple[int, str]:
    priority = clean(lead.get("prioridad")).lower()
    rank = 2
    if priority == "alta":
        rank = 0
    elif priority == "media":
        rank = 1
    return rank, clean(lead.get("sector"))


def row_key(values: dict[str, object]) -> tuple[str, str]:
    return clean(values.get("Empresa") or values.get("nombre_empresa")), clean(values.get("Telefono") or values.get("telefono"))


def normalize_state(value: object) -> str:
    state = clean(value)
    if state in {"", "Nuevo", "No contactado"}:
        return "Pendiente"
    return state if state in STATE_OPTIONS else "Pendiente"


def normalize_result(value: object) -> str:
    result = clean(value)
    return result if result in RESULT_OPTIONS else "Pendiente"


def read_existing_updates(path: Path) -> dict[tuple[str, str], dict[str, object]]:
    if not path.exists():
        return {}

    wb = load_workbook(path, data_only=False)
    if MAIN_SHEET not in wb.sheetnames:
        return {}

    ws = wb[MAIN_SHEET]
    headers = [clean(cell.value) for cell in ws[1]]
    updates: dict[tuple[str, str], dict[str, object]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        values = dict(zip(headers, row))
        key = row_key(values)
        if not all(key):
            continue
        row_updates = {column: values.get(column) for column in EDITABLE_COLUMNS if values.get(column) not in (None, "")}
        if row_updates.get("Llamar") not in (None, "", *YES_NO_OPTIONS):
            row_updates.pop("Llamar", None)
        if "Estado llamada" in row_updates:
            row_updates["Estado llamada"] = normalize_state(row_updates["Estado llamada"])
        if "Resultado" in row_updates:
            row_updates["Resultado"] = normalize_result(row_updates["Resultado"])
        updates[key] = row_updates

    return updates


def lead_to_row(index: int, lead: dict[str, object], updates: dict[tuple[str, str], dict[str, object]]) -> list[object]:
    contact = clean(lead.get("decisor_nombre")) or "Responsable"
    base = {
        "Orden": index,
        "Llamar": "Si",
        "Sector": clean(lead.get("sector")),
        "Empresa": clean(lead.get("nombre_empresa")),
        "Telefono": clean(lead.get("telefono")),
        "Contacto": contact,
        "Cargo": clean(lead.get("decisor_cargo")),
        "Prioridad": clean(lead.get("prioridad")),
        "Problema probable": clean(lead.get("problema_visible")),
        "Frase de entrada": clean(lead.get("mensaje_llamada_personalizado")),
        "Estado llamada": normalize_state(lead.get("estado")),
        "Resultado": "Pendiente",
        "Fecha cita": "",
        "Hora cita": "",
        "Fecha seguimiento": "",
        "Proximo paso": "",
        "Notas": "",
    }
    base.update(updates.get(row_key(base), {}))
    return [base[column] for column in CALL_COLUMNS]


def setup_sheet(ws, title: str) -> None:
    ws.title = title
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def style_header(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def style_base_rows(ws) -> None:
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def set_widths(ws) -> None:
    widths = {
        "A": 8,
        "B": 10,
        "C": 18,
        "D": 36,
        "E": 16,
        "F": 22,
        "G": 24,
        "H": 12,
        "I": 42,
        "J": 58,
        "K": 22,
        "L": 24,
        "M": 16,
        "N": 14,
        "O": 18,
        "P": 28,
        "Q": 42,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def add_validations(ws) -> None:
    max_row = max(ws.max_row, 2)
    validations = [
        (DataValidation(type="list", formula1="=OpcionesLlamar", allow_blank=False), f"B2:B{max_row}"),
        (DataValidation(type="list", formula1="=OpcionesEstado", allow_blank=True), f"K2:K{max_row}"),
        (DataValidation(type="list", formula1="=OpcionesResultado", allow_blank=True), f"L2:L{max_row}"),
        (DataValidation(type="date", operator="greaterThan", formula1="DATE(2026,1,1)", allow_blank=True), f"M2:M{max_row}"),
        (DataValidation(type="time", operator="between", formula1="TIME(8,0,0)", formula2="TIME(21,0,0)", allow_blank=True), f"N2:N{max_row}"),
        (DataValidation(type="date", operator="greaterThan", formula1="DATE(2026,1,1)", allow_blank=True), f"O2:O{max_row}"),
    ]
    for validation, cell_range in validations:
        ws.add_data_validation(validation)
        validation.add(cell_range)


def add_call_sheet(wb: Workbook, leads: list[dict[str, object]], updates: dict[tuple[str, str], dict[str, object]]) -> None:
    ws = wb.active
    setup_sheet(ws, MAIN_SHEET)
    ws.append(CALL_COLUMNS)
    for index, lead in enumerate(leads, start=1):
        ws.append(lead_to_row(index, lead, updates))

    style_header(ws)
    style_base_rows(ws)
    set_widths(ws)
    add_validations(ws)

    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 54


def add_options_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Opciones")
    setup_sheet(ws, "Opciones")
    ws.append(["Estado llamada", "Resultado", "Llamar"])
    max_len = max(len(STATE_OPTIONS), len(RESULT_OPTIONS), len(YES_NO_OPTIONS))
    for index in range(max_len):
        ws.append([
            STATE_OPTIONS[index] if index < len(STATE_OPTIONS) else "",
            RESULT_OPTIONS[index] if index < len(RESULT_OPTIONS) else "",
            YES_NO_OPTIONS[index] if index < len(YES_NO_OPTIONS) else "",
        ])
    style_header(ws)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 14

    quoted = quote_sheetname("Opciones")
    wb.defined_names.add(DefinedName("OpcionesEstado", attr_text=f"{quoted}!$A$2:$A${len(STATE_OPTIONS) + 1}"))
    wb.defined_names.add(DefinedName("OpcionesResultado", attr_text=f"{quoted}!$B$2:$B${len(RESULT_OPTIONS) + 1}"))
    wb.defined_names.add(DefinedName("OpcionesLlamar", attr_text=f"{quoted}!$C$2:$C${len(YES_NO_OPTIONS) + 1}"))


def add_summary_sheet(wb: Workbook, leads: list[dict[str, object]]) -> None:
    ws = wb.create_sheet("Resumen")
    setup_sheet(ws, "Resumen")
    counts = Counter(clean(lead.get("sector")) for lead in leads)
    rows = [
        ["Resumen llamadas martes", ""],
        ["Total leads", len(leads)],
        ["Objetivo minimo llamadas", 15],
        ["Objetivo bueno conversaciones", "3-5"],
        ["Objetivo excelente", "1 cita/diagnostico agendado"],
        ["Recordatorio", "No vendas. Agenda diagnostico."],
        ["", ""],
        ["Leads por sector", ""],
    ]
    rows.extend([[sector, total] for sector, total in counts.items()])
    for row in rows:
        ws.append(row)

    ws["A1"].fill = HEADER_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["B1"].fill = HEADER_FILL
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 44


def main() -> None:
    parser = argparse.ArgumentParser(description="Exporta un Excel formateado de llamadas.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--limite", type=int, default=25)
    args = parser.parse_args()

    updates = read_existing_updates(args.output)
    leads = read_leads(args.input, args.limite)
    wb = Workbook()
    add_call_sheet(wb, leads, updates)
    add_summary_sheet(wb, leads)
    add_options_sheet(wb)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Excel generado: {args.output}")


if __name__ == "__main__":
    main()
